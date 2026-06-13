import io
from dataclasses import dataclass
from datetime import datetime


# Column mapping: internal_field_name → Excel column header
# UPDATE THIS when the broker provides the actual Excel model.
EXPECTED_COLUMNS: dict[str, str] = {
    "cliente_nombre": "Cliente",
    "cliente_email": "Email",
    "cuenta_comitente": "Cuenta Comitente",
    "cuenta_cotapartista": "Cuenta Cotapartista",
    "instrumento": "Instrumento",
    "tipo": "Tipo Operación",
    "cantidad": "Cantidad",
    "precio": "Precio",
    "moneda": "Moneda",
    "liquidacion": "Liquidación",
    "fecha_operacion": "Fecha Operación",
}

VALID_TIPOS = {"COMPRA", "VENTA"}
VALID_LIQUIDACIONES = {"CI", "24HS", "48HS"}


@dataclass
class OrdenParsed:
    cliente_nombre: str
    cliente_email: str
    cuenta_comitente: str
    cuenta_cotapartista: str
    instrumento: str
    tipo: str
    cantidad: float
    precio: float
    moneda: str
    liquidacion: str
    fecha_operacion: datetime


@dataclass
class RowError:
    fila: int
    mensaje: str


@dataclass
class ParseResult:
    ordenes: list[OrdenParsed]
    errors: list[RowError]


def parse_excel_file(file_bytes: bytes) -> ParseResult:
    import openpyxl

    wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Read headers from row 1
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    headers = [h.strip() if isinstance(h, str) else h for h in headers]

    expected_headers = set(EXPECTED_COLUMNS.values())
    present_headers = {h for h in headers if h is not None}
    missing = expected_headers - present_headers
    if missing:
        raise ValueError(f"Columnas faltantes: {', '.join(sorted(missing))}")

    # Build column index: header name → 0-based list index
    header_to_col = {h: idx for idx, h in enumerate(headers)}
    # Build field name → 0-based list index
    field_to_col = {
        field: header_to_col[header]
        for field, header in EXPECTED_COLUMNS.items()
    }

    ordenes: list[OrdenParsed] = []
    errors: list[RowError] = []

    for row_idx in range(2, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]

        # Skip completely empty rows
        if all(v is None for v in row_values):
            continue

        def get(field: str, _rv=row_values) -> object:
            return _rv[field_to_col[field]]

        try:
            orden = _parse_row(row_idx, get)
            ordenes.append(orden)
        except ValueError as e:
            errors.append(RowError(fila=row_idx, mensaje=str(e)))

    return ParseResult(ordenes=ordenes, errors=errors)


def _parse_row(row_idx: int, get) -> OrdenParsed:
    cliente_nombre = str(get("cliente_nombre") or "").strip()
    if not cliente_nombre:
        raise ValueError("cliente_nombre es obligatorio")

    cliente_email = str(get("cliente_email") or "").strip()
    if "@" not in cliente_email:
        raise ValueError(f"cliente_email inválido: '{cliente_email}'")

    cuenta_comitente = str(get("cuenta_comitente") or "").strip()
    if not cuenta_comitente:
        raise ValueError("cuenta_comitente es obligatoria")

    cuenta_cotapartista = str(get("cuenta_cotapartista") or "").strip()
    if not cuenta_cotapartista:
        raise ValueError("cuenta_cotapartista es obligatoria")

    instrumento = str(get("instrumento") or "").strip()
    if not instrumento:
        raise ValueError("instrumento es obligatorio")

    tipo = str(get("tipo") or "").strip().upper()
    if tipo not in VALID_TIPOS:
        raise ValueError(f"tipo inválido: '{tipo}'. Esperado: {VALID_TIPOS}")

    try:
        cantidad = float(get("cantidad"))
    except (TypeError, ValueError):
        raise ValueError("cantidad debe ser un número")
    if cantidad <= 0:
        raise ValueError(f"cantidad debe ser mayor a 0, se recibió: {cantidad}")

    try:
        precio = float(get("precio"))
    except (TypeError, ValueError):
        raise ValueError("precio debe ser un número")
    if precio <= 0:
        raise ValueError(f"precio debe ser mayor a 0, se recibió: {precio}")

    moneda = str(get("moneda") or "").strip()
    if not moneda:
        raise ValueError("moneda es obligatoria")

    liquidacion = str(get("liquidacion") or "").strip().upper()
    if liquidacion not in VALID_LIQUIDACIONES:
        raise ValueError(f"liquidacion inválida: '{liquidacion}'. Esperado: {VALID_LIQUIDACIONES}")

    raw_fecha = get("fecha_operacion")
    if isinstance(raw_fecha, datetime):
        fecha_operacion = raw_fecha
    elif isinstance(raw_fecha, str):
        try:
            fecha_operacion = datetime.fromisoformat(raw_fecha)
        except ValueError:
            raise ValueError(f"fecha_operacion inválida: '{raw_fecha}'")
    else:
        raise ValueError("fecha_operacion inválida o faltante")

    return OrdenParsed(
        cliente_nombre=cliente_nombre,
        cliente_email=cliente_email,
        cuenta_comitente=cuenta_comitente,
        cuenta_cotapartista=cuenta_cotapartista,
        instrumento=instrumento,
        tipo=tipo,
        cantidad=cantidad,
        precio=precio,
        moneda=moneda,
        liquidacion=liquidacion,
        fecha_operacion=fecha_operacion,
    )
