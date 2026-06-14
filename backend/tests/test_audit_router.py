import pytest


def test_get_audit_trail_for_orden(client, auth_headers, seeded_borrador_orden):
    r = client.get(f"/audit/{seeded_borrador_orden}", headers=auth_headers)
    assert r.status_code == 200
    events = r.json()
    assert isinstance(events, list)
    assert len(events) >= 1
    # The upload creates a CREADA event
    acciones = [e["accion"] for e in events]
    assert "CREADA" in acciones


def test_audit_event_has_correct_fields(client, auth_headers, seeded_borrador_orden):
    r = client.get(f"/audit/{seeded_borrador_orden}", headers=auth_headers)
    event = r.json()[0]
    assert "id" in event
    assert "orden_id" in event
    assert "accion" in event
    assert "timestamp" in event
    assert event["orden_id"] == seeded_borrador_orden


def test_audit_trail_grows_with_actions(client, auth_headers, seeded_borrador_orden):
    # Approve the orden — should add an APROBADA event
    client.post(f"/orders/{seeded_borrador_orden}/approve", headers=auth_headers)
    r = client.get(f"/audit/{seeded_borrador_orden}", headers=auth_headers)
    acciones = [e["accion"] for e in r.json()]
    assert "CREADA" in acciones
    assert "APROBADA" in acciones


def test_audit_trail_unknown_orden_returns_empty(client, auth_headers):
    r = client.get("/audit/00000000-0000-0000-0000-000000000000", headers=auth_headers)
    # Unknown orden → empty list (not 404, since no event ≠ orden doesn't exist from audit's perspective)
    assert r.status_code == 200
    assert r.json() == []


def test_export_audit_excel(client, auth_headers, seeded_borrador_orden):
    r = client.get(f"/audit/{seeded_borrador_orden}/export/excel", headers=auth_headers)
    assert r.status_code == 200
    assert "attachment" in r.headers.get("content-disposition", "")
    assert len(r.content) > 100  # non-trivial xlsx bytes


def test_export_audit_excel_is_valid_xlsx(client, auth_headers, seeded_borrador_orden):
    import io
    import openpyxl
    r = client.get(f"/audit/{seeded_borrador_orden}/export/excel", headers=auth_headers)
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    assert "Timestamp" in headers
    assert "Acción" in headers
    assert ws.max_row >= 2  # at least one data row (CREADA event)


def test_audit_requires_auth(client, seeded_borrador_orden):
    r = client.get(f"/audit/{seeded_borrador_orden}")
    assert r.status_code == 403

    r = client.get(f"/audit/{seeded_borrador_orden}/export/excel")
    assert r.status_code == 403
