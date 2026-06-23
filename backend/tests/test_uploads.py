# backend/tests/test_uploads.py
import io
import pytest
import openpyxl
from app.services.excel_parser import EXPECTED_COLUMNS
from app.models.config_dj import ConfigDJ
import app.services.session_store as store


@pytest.fixture(autouse=True)
def _clean_config_dj(db):
    """Ensure config_dj table is empty before each test for isolation."""
    db.query(ConfigDJ).delete()
    db.commit()


def make_excel_bytes(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = list(EXPECTED_COLUMNS.values())
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(EXPECTED_COLUMNS.keys(), 1):
            ws.cell(row=row_idx, column=col_idx, value=row[key])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


VALID_ROW = {
    "cliente_nombre": "Ana García",
    "cuenta_comitente": "11111",
    "cuenta_cotapartista": "22222",
    "id_orden": 999001,
    "fecha": "14/06/2026",
    "hora": "10:30:00",
    "fecha_liquidacion": "16/06/2026",
    "operacion": "Compra CI",
    "instrumento": "AL30",
    "moneda": "Pesos",
    "cantidad": 100.0,
    "precio": 70.5,
    "monto": 7050.0,
    "estado": "Ejecutada",
    "cantidad_operada": 100.0,
    "precio_operado": 70.5,
    "operador": "testuser",
    "origen": "Cliente",
    "asesor": "Test Asesor",
    "requiere_conformidad": 0,
}


def test_upload_returns_201_and_minutas(client, auth_headers):
    excel = make_excel_bytes([VALID_ROW])
    r = client.post(
        "/uploads/excel",
        files={"file": ("ops.xlsx", excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_headers,
    )
    assert r.status_code == 201
    data = r.json()
    assert data["ordenes_validas"] == 1
    assert data["ordenes_con_error"] == 0
    assert len(data["minutas"]) == 1
    minuta = data["minutas"][0]
    assert minuta["cliente_nombre"] == "Ana García"
    assert minuta["estado"] == "BORRADOR"
    assert minuta["texto_minuta"] != ""
    assert "id" in minuta


def test_upload_two_rows_creates_two_minutas(client, auth_headers):
    row2 = {**VALID_ROW, "cliente_nombre": "Pedro López", "id_orden": 999002}
    excel = make_excel_bytes([VALID_ROW, row2])
    r = client.post(
        "/uploads/excel",
        files={"file": ("ops.xlsx", excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_headers,
    )
    assert r.status_code == 201
    assert r.json()["ordenes_validas"] == 2
    assert len(r.json()["minutas"]) == 2


def test_upload_bad_extension_returns_400(client, auth_headers):
    r = client.post(
        "/uploads/excel",
        files={"file": ("ops.csv", b"a,b,c", "text/csv")},
        headers=auth_headers,
    )
    assert r.status_code == 400


def test_upload_missing_column_returns_400(client, auth_headers):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Columna Incorrecta")
    ws.cell(row=2, column=1, value="valor")
    buf = io.BytesIO()
    wb.save(buf)
    r = client.post(
        "/uploads/excel",
        files={"file": ("ops.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_headers,
    )
    assert r.status_code == 400


def test_upload_requires_auth(client):
    excel = make_excel_bytes([VALID_ROW])
    r = client.post(
        "/uploads/excel",
        files={"file": ("ops.xlsx", excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 403


def test_upload_ordenes_filtradas_count(client, auth_headers):
    """Filas que no coinciden con ningún filtro activo no se filtran (sin reglas configuradas)."""
    excel = make_excel_bytes([VALID_ROW])
    r = client.post(
        "/uploads/excel",
        files={"file": ("ops.xlsx", excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_headers,
    )
    assert r.status_code == 201
    data = r.json()
    assert "ordenes_filtradas" in data
    assert data["ordenes_filtradas"] == 0


def test_upload_with_multiple_djs(client, auth_headers, db):
    """When multiple DJs are active, all should be evaluated and matching texts concatenated."""
    from app.services.db_config import create_config_dj, ConfigDJData

    create_config_dj(db, ConfigDJData(
        nombre="DJ monto alto",
        activa=True,
        incluir_texto_en_minuta=True,
        texto_alerta="Alerta por monto alto",
        reglas=[{"campo": "monto", "operador": ">=", "valor": "1000"}],
        logica="OR",
    ))
    create_config_dj(db, ConfigDJData(
        nombre="DJ operacion compra",
        activa=True,
        incluir_texto_en_minuta=True,
        texto_alerta="Alerta por compra",
        reglas=[{"campo": "operacion", "operador": "=", "valor": "Compra CI"}],
        logica="OR",
    ))

    excel = make_excel_bytes([VALID_ROW])
    r = client.post(
        "/uploads/excel",
        files={"file": ("ops.xlsx", excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_headers,
    )
    assert r.status_code == 201
    minuta = r.json()["minutas"][0]
    assert minuta["dj_aplicada"] is True
    assert "Alerta por monto alto" in minuta["dj_texto"]
    assert "Alerta por compra" in minuta["dj_texto"]


def test_upload_inactive_dj_ignored(client, auth_headers, db):
    """DJs with activa=False should not be evaluated."""
    from app.services.db_config import create_config_dj, ConfigDJData

    create_config_dj(db, ConfigDJData(
        nombre="DJ inactiva",
        activa=False,
        incluir_texto_en_minuta=True,
        texto_alerta="No debería aparecer",
        reglas=[{"campo": "monto", "operador": ">=", "valor": "1"}],
        logica="OR",
    ))

    excel = make_excel_bytes([VALID_ROW])
    r = client.post(
        "/uploads/excel",
        files={"file": ("ops.xlsx", excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_headers,
    )
    assert r.status_code == 201
    minuta = r.json()["minutas"][0]
    assert minuta["dj_aplicada"] is False


def test_partial_errors_reported_without_blocking(client, auth_headers):
    bad_row = {**VALID_ROW, "id_orden": "NO_ES_NUMERO", "cantidad": "NO_ES_NUMERO"}
    excel = make_excel_bytes([VALID_ROW, bad_row])
    r = client.post(
        "/uploads/excel",
        files={"file": ("ops.xlsx", excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_headers,
    )
    assert r.status_code == 201
    data = r.json()
    assert data["ordenes_validas"] == 1
    assert data["ordenes_con_error"] == 1
    assert len(data["errors"]) == 1
    assert len(data["minutas"]) == 1
