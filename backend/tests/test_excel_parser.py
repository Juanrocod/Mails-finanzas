import io
import pytest
import openpyxl
from datetime import datetime

from app.services.excel_parser import parse_excel_file, EXPECTED_COLUMNS


def make_excel(rows: list[dict]) -> bytes:
    """Build an in-memory .xlsx with the correct headers and the given rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = list(EXPECTED_COLUMNS.values())
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, key in enumerate(EXPECTED_COLUMNS.keys(), 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(key))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


VALID_ROW = {
    "cliente_nombre": "Juan Pérez",
    "cliente_email": "juan@example.com",
    "cuenta_comitente": "12345",
    "cuenta_cotapartista": "67890",
    "instrumento": "AL30",
    "tipo": "COMPRA",
    "cantidad": 1000.0,
    "precio": 70.50,
    "moneda": "USD",
    "liquidacion": "24HS",
    "fecha_operacion": datetime(2026, 6, 13, 10, 30),
}


def test_parse_valid_single_row():
    result = parse_excel_file(make_excel([VALID_ROW]))
    assert len(result.ordenes) == 1
    assert len(result.errors) == 0
    assert result.ordenes[0].instrumento == "AL30"
    assert result.ordenes[0].cliente_email == "juan@example.com"
    assert result.ordenes[0].tipo == "COMPRA"


def test_parse_missing_column_raises():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "SoloUnaColumna"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="Columnas faltantes"):
        parse_excel_file(buf.getvalue())


def test_parse_invalid_tipo_reports_error():
    bad_row = {**VALID_ROW, "tipo": "INVALIDO"}
    result = parse_excel_file(make_excel([VALID_ROW, bad_row]))
    assert len(result.ordenes) == 1
    assert len(result.errors) == 1
    assert result.errors[0].fila == 3  # row 2 = first data row, row 3 = second


def test_parse_zero_cantidad_reports_error():
    result = parse_excel_file(make_excel([{**VALID_ROW, "cantidad": 0}]))
    assert len(result.ordenes) == 0
    assert len(result.errors) == 1


def test_parse_negative_precio_reports_error():
    result = parse_excel_file(make_excel([{**VALID_ROW, "precio": -1}]))
    assert len(result.ordenes) == 0
    assert len(result.errors) == 1


def test_parse_invalid_email_reports_error():
    result = parse_excel_file(make_excel([{**VALID_ROW, "cliente_email": "notanemail"}]))
    assert len(result.ordenes) == 0
    assert len(result.errors) == 1


def test_parse_invalid_liquidacion_reports_error():
    result = parse_excel_file(make_excel([{**VALID_ROW, "liquidacion": "99HS"}]))
    assert len(result.ordenes) == 0
    assert len(result.errors) == 1


def test_parse_empty_nombre_reports_error():
    result = parse_excel_file(make_excel([{**VALID_ROW, "cliente_nombre": ""}]))
    assert len(result.ordenes) == 0
    assert len(result.errors) == 1


def test_parse_multiple_rows_independent_errors():
    """Error on one row does not block other rows."""
    rows = [
        VALID_ROW,
        {**VALID_ROW, "tipo": "INVALIDO"},
        VALID_ROW,
        {**VALID_ROW, "cantidad": 0},
    ]
    result = parse_excel_file(make_excel(rows))
    assert len(result.ordenes) == 2
    assert len(result.errors) == 2


def test_parse_empty_excel_no_data():
    result = parse_excel_file(make_excel([]))
    assert len(result.ordenes) == 0
    assert len(result.errors) == 0
